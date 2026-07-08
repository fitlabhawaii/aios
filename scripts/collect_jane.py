"""
DataOS — Jane App Sales Collector

Jane has no public API, so we ingest the Sales report that you export from Jane
(Reports -> Sales -> Export). Drop the .xlsx (or .csv) file into:

    data/imports/jane/

This collector reads the MOST RECENT file in that folder, parses every line item,
and upserts it into the `jane_sales` table. Because Jane exports the full history
each time and every row has a unique Invoice #, re-importing a fresh export simply
updates existing rows and adds new ones — safe to run repeatedly.

NOTE: This data contains patient names (PHI). It lives only in your local database
(data/data.db) and the import folder — both are gitignored and never leave your Mac.

Tables created: jane_sales
"""

import glob
import os
from datetime import datetime, timezone
from pathlib import Path

WORKSPACE_ROOT = Path(__file__).resolve().parent.parent
IMPORT_DIR = WORKSPACE_ROOT / "data" / "imports" / "jane"

# Maps Jane's export column headers -> our database column names
COLUMN_MAP = {
    "Location": "location",
    "Purchase Date": "purchase_date",
    "Invoice Date": "invoice_date",
    "Patient Guid": "patient_guid",
    "Patient": "patient",
    "Item": "item",
    "Staff Member": "staff_member",
    "Payer": "payer",
    "Invoice #": "invoice_number",
    "Income Category": "income_category",
    "Details": "details",
    "Status": "status",
    "Subtotal": "subtotal",
    "Sales Tax": "sales_tax",
    "Total": "total",
    "Collected": "collected",
    "Balance": "balance",
}

NUMERIC_COLS = {"subtotal", "sales_tax", "total", "collected", "balance"}


def _latest_export():
    """Return the newest Jane export file in the import folder, or None."""
    files = []
    for pattern in ("*.xlsx", "*.csv"):
        files.extend(glob.glob(str(IMPORT_DIR / pattern)))
    if not files:
        return None
    return max(files, key=os.path.getmtime)


def _to_day(value):
    """Extract a YYYY-MM-DD day string from a datetime or a date-ish string."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    # Handles "2024-05-05 09:36:50 -1000", "2024-05-05T09:36:50", "2024-05-05", etc.
    for sep in (" ", "T"):
        if sep in text:
            text = text.split(sep)[0]
            break
    return text[:10] if len(text) >= 10 else text


def _to_num(value):
    """Coerce a cell to float; return 0.0 if blank/non-numeric."""
    if value is None or value == "":
        return 0.0
    try:
        return float(str(value).replace("$", "").replace(",", "").strip())
    except (ValueError, AttributeError):
        return 0.0


def _read_rows(filepath):
    """Read the export into a list of dicts keyed by our db column names."""
    path = Path(filepath)
    records = []

    if path.suffix.lower() == ".xlsx":
        try:
            import openpyxl
        except ImportError:
            raise ImportError("Missing 'openpyxl' — run: .venv/bin/pip install openpyxl")
        wb = openpyxl.load_workbook(str(path), data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(rows)]
        for raw in rows:
            if raw is None or all(c is None for c in raw):
                continue
            records.append(dict(zip(headers, raw)))
    else:  # CSV fallback
        import csv
        with open(path, newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                records.append(row)

    # Normalize to our column names
    out = []
    for r in records:
        rec = {}
        for jane_col, db_col in COLUMN_MAP.items():
            rec[db_col] = r.get(jane_col)
        if not rec.get("invoice_number"):
            continue  # skip rows without an invoice id
        rec["purchase_day"] = _to_day(rec.get("purchase_date"))
        for c in NUMERIC_COLS:
            rec[c] = _to_num(rec.get(c))
        rec["purchase_date"] = str(rec.get("purchase_date")) if rec.get("purchase_date") else None
        rec["invoice_date"] = str(rec.get("invoice_date")) if rec.get("invoice_date") else None
        out.append(rec)
    return out


def collect():
    """Read the latest Jane export from data/imports/jane/."""
    filepath = _latest_export()
    if not filepath:
        return {
            "source": "jane",
            "status": "skipped",
            "reason": "No export found in data/imports/jane/ — export a Sales report from "
                      "Jane (Reports -> Sales -> Export) and drop the file there.",
        }
    try:
        rows = _read_rows(filepath)
        return {
            "source": "jane",
            "status": "success",
            "data": {"rows": rows, "file": os.path.basename(filepath)},
        }
    except Exception as e:
        return {"source": "jane", "status": "error", "reason": str(e)}


def write(conn, result, date):
    """Upsert Jane sales line items. Returns records written."""
    conn.execute("""
        CREATE TABLE IF NOT EXISTS jane_sales (
            invoice_number  TEXT PRIMARY KEY,
            date            TEXT,           -- purchase day (YYYY-MM-DD), for freshness/grouping
            location        TEXT,
            purchase_date   TEXT,           -- full timestamp as exported
            invoice_date    TEXT,
            patient_guid    TEXT,
            patient         TEXT,           -- PHI — local only
            item            TEXT,
            staff_member    TEXT,
            payer           TEXT,
            income_category TEXT,
            details         TEXT,
            status          TEXT,
            subtotal        REAL,
            sales_tax       REAL,
            total           REAL,
            collected       REAL,
            balance         REAL,
            imported_at     TEXT
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_jane_day ON jane_sales(date)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_jane_patient ON jane_sales(patient_guid)")

    if result.get("status") != "success":
        conn.commit()
        return 0

    imported_at = datetime.now(timezone.utc).isoformat()
    rows = result["data"]["rows"]
    n = 0
    for r in rows:
        conn.execute("""
            INSERT OR REPLACE INTO jane_sales (
                invoice_number, date, location, purchase_date, invoice_date,
                patient_guid, patient, item, staff_member, payer, income_category,
                details, status, subtotal, sales_tax, total, collected, balance, imported_at
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            r["invoice_number"], r["purchase_day"], r.get("location"), r.get("purchase_date"),
            r.get("invoice_date"), r.get("patient_guid"), r.get("patient"), r.get("item"),
            r.get("staff_member"), r.get("payer"), r.get("income_category"), r.get("details"),
            r.get("status"), r["subtotal"], r["sales_tax"], r["total"], r["collected"],
            r["balance"], imported_at,
        ))
        n += 1
    conn.commit()
    return n


if __name__ == "__main__":
    res = collect()
    if res["status"] == "success":
        rows = res["data"]["rows"]
        total = sum(r["collected"] for r in rows)
        days = [r["purchase_day"] for r in rows if r["purchase_day"]]
        print(f"Parsed {len(rows)} line items from {res['data']['file']}")
        print(f"Date range: {min(days)} -> {max(days)}")
        print(f"Total collected: ${total:,.2f}")
    else:
        print(f"{res['status']}: {res.get('reason')}")
